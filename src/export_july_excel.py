"""Export the July 2026 backtest to a formatted Excel workbook."""
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import config

OUT = config.RESULTS_DIR / "july2026_backtest.xlsx"

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
SECTION_FONT = Font(bold=True, size=11, color="1F4E79")
GREEN = Font(color="006100")
RED = Font(color="9C0006")
THIN = Border(*[Side(style="thin", color="D9D9D9")] * 4)
INR = '#,##0"  "'
INR2 = '#,##0.00'


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN


def autofit(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def pnl_font(cell, value):
    cell.font = GREEN if value > 0 else RED if value < 0 else Font()


def main():
    tr = pd.read_csv(config.RESULTS_DIR / "trades_july2026.csv",
                     parse_dates=["entry_ts", "exit_ts"])
    base = pd.read_csv(config.RESULTS_DIR / "trades_baseline.csv")
    sig = pd.read_parquet(config.RESULTS_DIR / "signals_july2026.parquet")

    tr["exit_price_spread"] = tr["exit_cost_to_close"]
    tr["hold_minutes"] = (tr["exit_ts"] - tr["entry_ts"]).dt.total_seconds() / 60
    tr["credit_rupees"] = tr["credit"] * tr["lot"]
    tr["max_loss_rupees"] = (config.WING_POINTS - tr["credit"]) * tr["lot"]
    tr["ret_on_margin_pct"] = tr["net_pnl"] / tr["margin"] * 100

    wb = Workbook()

    # ---------------- Sheet 1: Summary ----------------
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws["B2"] = "NIFTY Credit Spread Strategy — July 2026 Backtest"
    ws["B2"].font = TITLE_FONT
    d0 = pd.Timestamp(tr["date"].min()).strftime("%d-%b-%Y")
    d1 = pd.Timestamp(tr["date"].max()).strftime("%d-%b-%Y")
    ws["B3"] = (f"Period: {d0} to {d1}  |  1 lot (lot size {int(tr['lot'].iloc[0])})  |  "
                "weekly expiry  |  all Indian F&O charges included")
    ws["B3"].font = Font(italic=True, color="595959")

    wins = tr[tr["net_pnl"] > 0]
    losses = tr[tr["net_pnl"] <= 0]
    daily = tr.groupby(tr["exit_ts"].dt.date)["net_pnl"].sum()
    eq = daily.cumsum()
    dd = (eq - eq.cummax()).min()
    sharpe = daily.mean() / daily.std() * np.sqrt(252) if daily.std() > 0 else np.nan

    rows = [
        ("Strategy rules", None, "section"),
        ("Entry window", "10:15 – 14:15 (signal: alpha & alpha2 cross 0.8 / 0.2)", None),
        ("Bullish trade", "Bull put spread: sell ATM PE, buy PE 400 pts lower", None),
        ("Bearish trade", "Bear call spread: sell ATM CE, buy CE 400 pts higher", None),
        ("Stop-loss", f"MTM loss ≥ {config.SL_PCT_OF_MARGIN:.0%} of margin (margin = width − credit)", None),
        ("Square-off", "15:15 same day (no overnight positions)", None),
        ("Performance", None, "section"),
        ("Trading days", len(daily), "int"),
        ("Trades", f"{len(tr)}  ({(tr['type']=='bull_put').sum()} bull put, {(tr['type']=='bear_call').sum()} bear call)", None),
        ("Win rate", len(wins) / len(tr), "pct"),
        ("Gross P&L", tr["gross_pnl"].sum(), "inr"),
        ("Total charges", tr["charges"].sum(), "inr"),
        ("Net P&L", tr["net_pnl"].sum(), "inr"),
        ("Average win", wins["net_pnl"].mean(), "inr"),
        ("Average loss", losses["net_pnl"].mean(), "inr"),
        ("Profit factor", wins["net_pnl"].sum() / max(1e-9, -losses["net_pnl"].sum()), "num"),
        ("Max drawdown (daily eq.)", dd, "inr"),
        ("Best day", daily.max(), "inr"),
        ("Worst day", daily.min(), "inr"),
        ("Daily Sharpe (annualised)", sharpe, "num"),
        ("Avg margin per trade", tr["margin"].mean(), "inr"),
        ("Return on avg margin", tr["net_pnl"].sum() / tr["margin"].mean(), "pct"),
        ("Stop-loss exits", (tr["exit_reason"] == "SL").sum(), "int"),
        ("Reality check (full history)", None, "section"),
        ("Baseline period", f"{base['date'].min()} to {base['date'].max()}", None),
        ("Baseline trades", len(base), "int"),
        ("Baseline win rate", (base["net_pnl"] > 0).mean(), "pct"),
        ("Baseline net P&L", base["net_pnl"].sum(), "inr"),
        ("Note", "July 2026 was a profitable month, but the strategy loses money "
                 "over the full 21-month history — treat this month as variance, "
                 "not validation.", None),
    ]
    r = 5
    for label, val, kind in rows:
        if kind == "section":
            ws.cell(row=r, column=2, value=label).font = SECTION_FONT
            r += 1
            continue
        ws.cell(row=r, column=2, value=label).font = Font(bold=True)
        c = ws.cell(row=r, column=3, value=val)
        if kind == "inr":
            c.number_format = '₹#,##0'
            pnl_font(c, val)
        elif kind == "pct":
            c.number_format = "0.0%"
        elif kind == "num":
            c.number_format = "0.00"
        c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    autofit(ws, [2, 30, 72])

    # ---------------- Sheet 2: Trades ----------------
    ws = wb.create_sheet("Trades")
    cols = [
        ("Date", "date", 12), ("Expiry", "expiry", 12), ("Type", "type", 11),
        ("Entry time", "entry_ts", 17), ("Exit time", "exit_ts", 17),
        ("Hold (min)", "hold_minutes", 10), ("Exit reason", "exit_reason", 11),
        ("Spot at entry", "spot_entry", 12), ("ATM", "atm", 9),
        ("Short strike", "short_strike", 11), ("Long strike", "long_strike", 11),
        ("Lot size", "lot", 9), ("Alpha", "alpha", 9), ("Alpha2", "alpha2", 9),
        ("Credit (pts)", "credit", 11), ("Credit (₹)", "credit_rupees", 12),
        ("Exit cost (pts)", "exit_price_spread", 12), ("Margin (₹)", "margin", 12),
        ("Max loss (₹)", "max_loss_rupees", 12), ("Gross P&L (₹)", "gross_pnl", 13),
        ("Charges (₹)", "charges", 11), ("Net P&L (₹)", "net_pnl", 13),
        ("Return on margin", "ret_on_margin_pct", 13),
    ]
    for j, (h, _, _) in enumerate(cols, 1):
        ws.cell(row=1, column=j, value=h)
    style_header(ws, 1, len(cols))
    for i, (_, row) in enumerate(tr.iterrows(), 2):
        for j, (_, field, _) in enumerate(cols, 1):
            v = row[field]
            if field in ("entry_ts", "exit_ts"):
                v = row[field].strftime("%d-%b %H:%M")
            c = ws.cell(row=i, column=j, value=v)
            c.border = THIN
            if field in ("gross_pnl", "net_pnl"):
                c.number_format = '₹#,##0'
                pnl_font(c, v)
            elif field in ("credit_rupees", "margin", "max_loss_rupees", "charges"):
                c.number_format = '₹#,##0'
            elif field in ("credit", "exit_price_spread", "spot_entry"):
                c.number_format = INR2
            elif field in ("alpha", "alpha2"):
                c.number_format = "0.000"
            elif field == "ret_on_margin_pct":
                c.number_format = '0.0"%"'
                pnl_font(c, v)
            elif field == "hold_minutes":
                c.number_format = "0"
    total_row = len(tr) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    for field, col_idx in (("gross_pnl", 20), ("charges", 21), ("net_pnl", 22)):
        c = ws.cell(row=total_row, column=col_idx, value=tr[field].sum())
        c.number_format = '₹#,##0'
        c.font = Font(bold=True)
        pnl_font(c, tr[field].sum())
    ws.freeze_panes = "A2"
    autofit(ws, [w for _, _, w in cols])

    # ---------------- Sheet 3: Daily P&L ----------------
    ws = wb.create_sheet("Daily P&L")
    headers = ["Date", "Trades", "Type", "Net P&L (₹)", "Cumulative (₹)", "Drawdown (₹)"]
    for j, h in enumerate(headers, 1):
        ws.cell(row=1, column=j, value=h)
    style_header(ws, 1, len(headers))
    cum = 0.0
    peak = 0.0
    for i, (d, grp) in enumerate(tr.groupby(tr["exit_ts"].dt.date), 2):
        pnl = grp["net_pnl"].sum()
        cum += pnl
        peak = max(peak, cum)
        vals = [str(d), len(grp), ", ".join(grp["type"]), pnl, cum, cum - peak]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=i, column=j, value=v)
            c.border = THIN
            if j >= 4:
                c.number_format = '₹#,##0'
                pnl_font(c, v)
    ws.freeze_panes = "A2"
    autofit(ws, [12, 8, 14, 14, 14, 14])

    # ---------------- Sheet 4: Charges breakdown ----------------
    ws = wb.create_sheet("Charges Detail")
    ws["B2"] = "Transaction cost model (per trade = 4 orders: 2 entry + 2 exit)"
    ws["B2"].font = SECTION_FONT
    items = [
        ("Brokerage", f"₹{config.BROKERAGE_PER_ORDER:.0f} per order (₹80 per trade)"),
        ("STT", "0.1% of sell-side premium turnover"),
        ("Exchange txn charge", "0.03503% of premium turnover"),
        ("SEBI charges", "0.0001% of turnover"),
        ("Stamp duty", "0.003% of buy-side turnover"),
        ("GST", "18% on brokerage + exchange + SEBI"),
        ("Slippage", f"₹{config.SLIPPAGE:.2f} per leg per fill (1 tick), built into fill prices"),
        ("", ""),
        ("Total charges, July", f"₹{tr['charges'].sum():,.0f} across {len(tr)} trades "
                                f"(avg ₹{tr['charges'].mean():,.0f}/trade)"),
    ]
    for i, (k, v) in enumerate(items, 4):
        ws.cell(row=i, column=2, value=k).font = Font(bold=True)
        ws.cell(row=i, column=3, value=v)
    autofit(ws, [2, 24, 70])

    wb.save(OUT)
    print(f"saved -> {OUT}")
    print("sheets:", wb.sheetnames)


if __name__ == "__main__":
    main()
